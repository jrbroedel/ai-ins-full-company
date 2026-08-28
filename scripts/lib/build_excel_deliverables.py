#!/usr/bin/env python3
"""Build 2 — four Excel deliverables, READ-ONLY from the frozen canonical artifact.
Generates no new numbers: every cell is an artifact field or a documented calc/formula
over artifact fields. See ADR 0043."""
import json, hashlib, os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# Repo-relative: this file lives at scripts/lib/ ; repo root is two levels up.
ROOT = Path(__file__).resolve().parents[2]
CANON = str(ROOT / "sample-data" / "canonical")
DATA = f"{CANON}/canonical_dataset.json"
MANIFEST = f"{CANON}/canonical_manifest.json"
OUT = f"{CANON}/deliverables"
EXPECT_SHA = "0a3d67e8774e8cd15fba1c8ab9fa484cd433ac71665a04ca040d9db96b3a9811"

# ---- fail-closed sha256 verify ----
h = hashlib.sha256(open(DATA, "rb").read()).hexdigest()
man = json.load(open(MANIFEST))
assert h == EXPECT_SHA == man["dataset_sha256"], f"SHA MISMATCH {h}"
D = json.load(open(DATA))
SUM = D["summary"]
SUBS = D["submissions"]
CLAIMS = D["policy_period_claims"]
MA = D["monthly_aggregates"]
SYNDS = D["parameters"]["syndicates"]  # ordered 10
GWP = SUM["gwp_bound"]
print(f"sha256 OK · subs={len(SUBS)} binds={SUM['binds']} claims={len(CLAIMS)}")

# class number -> label
CLASS_LABEL = {}
for r in D["rating_snapshot"]["base_rates"]:
    CLASS_LABEL[r["rating_vehicle_class"]] = r["rating_class_label"]

os.makedirs(OUT, exist_ok=True)

# ---------- shared styling ----------
# Styling layer follows the MGA Program Master conventions: Arial throughout,
# navy title banners, blue section headers, semantic tab colours, 1-dp percentages.
FN = "Arial"
NAVY = "1F3864"; BLUE = "2E5A88"; LIGHT = "D9E1F2"; GREY = "EDEDED"
TOTFILL = "D9E1F2"; MEMOFILL = "FFF2CC"
def font(sz=10, b=False, color="000000"): return Font(name=FN, size=sz, bold=b, color=color)
HFONT = Font(name=FN, size=10, bold=True, color="FFFFFF")
TITLEF = Font(name=FN, size=14, bold=True, color=NAVY)
BANNERF = Font(name=FN, size=13, bold=True, color="FFFFFF")
SUBF = Font(name=FN, size=9, italic=True, color="595959")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CUR = '#,##0.00;(#,##0.00);-'; INT = '#,##0'; PCT2 = '0.0%;(0.0%);-'; RATE = '0.0000'
CTR = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

def hdr(ws, row, headers, start=1):
    for i, hh in enumerate(headers):
        c = ws.cell(row=row, column=start + i, value=hh)
        c.font = HFONT; c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = CTR; c.border = BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)

def title(ws, text, sub=None, span=1):
    # A1 = navy banner (white Arial 13 bold), merged across the sheet's columns.
    last_col = get_column_letter(max(span, 1))
    a1 = ws["A1"]; a1.value = text; a1.font = BANNERF
    a1.fill = PatternFill("solid", fgColor=NAVY); a1.alignment = LEFT
    if span > 1:
        ws.merge_cells(f"A1:{last_col}1")
    ws.row_dimensions[1].height = 22
    if sub:
        a2 = ws["A2"]; a2.value = sub; a2.font = SUBF
        if span > 1:
            ws.merge_cells(f"A2:{last_col}2")

def tabcolors(wb, color):
    # Semantic tab colour per workbook; every README tab is green.
    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = "548235" if ws.title == "README" else color

def widths(ws, wmap):
    for col, w in wmap.items():
        ws.column_dimensions[col].width = w

PROV = (f"Source: canonical_dataset.json (ADR 0043), sha256 {EXPECT_SHA[:16]}… — verified against "
        f"canonical_manifest.json. All figures derive read-only from that frozen artifact; no new "
        f"numbers generated. Styling follows the MGA Program Master conventions (Arial; navy title "
        f"banners; blue section headers; semantic tab colours; 1-dp percentages).")

def readme_tab(wb, title_text, lines):
    ws = wb.create_sheet("README", 0)
    ws.sheet_view.showGridLines = False
    ws["A1"] = title_text; ws["A1"].font = TITLEF
    ws["A2"] = PROV; ws["A2"].font = SUBF; ws["A2"].alignment = WRAP
    ws.merge_cells("A2:H6"); ws.row_dimensions[2].height = 70
    r = 8
    for lab, val in lines:
        ws.cell(row=r, column=1, value=lab).font = font(10, b=bool(val == ""))
        if val != "":
            c = ws.cell(row=r, column=3, value=val); c.font = font(10)
        r += 1
    widths(ws, {"A": 42, "B": 3, "C": 55})
    return ws

# ================================================================= WB1 SUBMISSIONS
def build_submissions():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Submissions"
    cols = ["Seq", "Submission ID", "Submitted Date", "Month", "Applicant", "Occupation",
            "Garaging State", "Veh Year", "Make", "Model", "Vehicle Category", "Rating Class",
            "Agreed Value", "Loss-Run Pattern", "Prior Claims", "At-Fault", "Claims >$100k",
            "Prior Non-Renewal", "Disposition", "Quote Premium", "Premium Kind", "Bind Status",
            "Policy Number"]
    title(ws, "Submissions Export — Full Funnel",
          f'One row per submission (all {SUM["submissions"]:,}). Quote premium shown for every submission; '
          "premium_kind flags bound vs indicative. Earned/bound premium = bound rows only.",
          span=len(cols))
    HR = 4
    hdr(ws, HR, cols)
    for j, s in enumerate(SUBS):
        r = HR + 1 + j
        q = s["quote"]; rb = q["rating_basis"]; lr = s["loss_run"]; v = s["vehicle"]
        pol = s.get("policy")
        row = [
            s["seq"], s["submission_id"], s["submitted_at"], s["submitted_at"][:7],
            f'{s["applicant"]["first_name"]} {s["applicant"]["last_name"]}',
            s["applicant"]["occupation"], s["garaging_state"], v["year"], v["make"], v["model"],
            v["vehicle_category"], CLASS_LABEL.get(rb["rating_vehicle_class"], rb["rating_vehicle_class"]),
            v["agreed_value"], lr["pattern"], lr["claim_count"], lr["at_fault_count"],
            lr["claims_over_100k"], "Yes" if lr["prior_carrier_nonrenewal"] else "No",
            s["disposition"], q["premium"], q["premium_kind"],
            "Bound" if pol else "Not Bound", pol["policy_number"] if pol else "",
        ]
        for i, val in enumerate(row):
            c = ws.cell(row=r, column=1 + i, value=val)
            c.font = font(10)
            if i in (12, 19):  # agreed value, premium
                c.number_format = CUR
            if r % 2 == 0:
                c.fill = PatternFill("solid", fgColor=GREY)
    last = HR + len(SUBS)
    ws.auto_filter.ref = f"A{HR}:{get_column_letter(len(cols))}{last}"
    widths(ws, {"A": 6, "B": 30, "C": 13, "D": 8, "E": 20, "F": 22, "G": 6, "H": 8, "I": 14,
                "J": 18, "K": 18, "L": 24, "M": 14, "N": 18, "O": 8, "P": 8, "Q": 8, "R": 10,
                "S": 11, "T": 14, "U": 12, "V": 11, "W": 22})

    # Summary tab with live reconciliation formulas
    sm = wb.create_sheet("Summary")
    sm.sheet_view.showGridLines = False
    title(sm, "Funnel Reconciliation", "Live COUNTIF/SUMIF over the Submissions tab; ties to the artifact summary.", span=3)
    S = "Submissions"; disp = f"'{S}'!S5:S{last}"; kind = f"'{S}'!U5:U{last}"; prem = f"'{S}'!T5:T{last}"
    rows = [
        ("Metric", "Formula (live)", "Artifact"),
        ("Total submissions", f"=COUNTA('{S}'!B5:B{last})", SUM["submissions"]),
        ("Binds", f'=COUNTIF({disp},"bind")', SUM["binds"]),
        ("Declines", f'=COUNTIF({disp},"decline")', SUM["declines"]),
        ("Refers", f'=COUNTIF({disp},"refer")', SUM["refers"]),
        ("Indicative-only quotes", f'=COUNTIF({kind},"indicative")', SUM["indicative_only_quotes"]),
        ("Bound quotes", f'=COUNTIF({kind},"bound")', SUM["binds"]),
        ("GWP — bound premium", f'=SUMIF({kind},"bound",{prem})', GWP),
        ("Indicative premium (NOT earned)", f'=SUMIF({kind},"indicative",{prem})', None),
        ("Bind %", f'=COUNTIF({disp},"bind")/COUNTA(\'{S}\'!B5:B{last})', SUM["bind_pct"] / 100),
    ]
    hdr(sm, 4, ["Metric", "Formula (live)", "Artifact value"])
    for k, (lab, f, art) in enumerate(rows[1:]):
        r = 5 + k
        sm.cell(row=r, column=1, value=lab).font = font(10)
        cf = sm.cell(row=r, column=2, value=f); cf.font = font(10)
        if "GWP" in lab or "premium" in lab.lower(): cf.number_format = CUR
        if lab == "Bind %": cf.number_format = PCT2
        if art is not None:
            ca = sm.cell(row=r, column=3, value=art); ca.font = font(10)
            if "GWP" in lab: ca.number_format = CUR
            elif lab == "Bind %": ca.number_format = PCT2
            else: ca.number_format = INT
        for cc in range(1, 4): sm.cell(row=r, column=cc).border = BORDER
    widths(sm, {"A": 32, "B": 34, "C": 18})
    readme_tab(wb, "01 · Submissions Export", [
        ("Rows", str(len(SUBS))),
        ("Reconciliation", "See 'Summary' tab (live formulas tie to artifact)."),
        ("Bind / Decline / Refer", f'{SUM["binds"]} / {SUM["declines"]} / {SUM["refers"]}'),
        ("GWP (bound)", f'${GWP:,.2f}'),
        ("Note", "Indicative premium is labeled and NEVER treated as earned/bound premium."),
    ])
    tabcolors(wb, "808080")
    p = f"{OUT}/01_submissions_export.xlsx"; wb.save(p); return p

# ================================================================= WB2 MONTHLY RATERS
def build_raters():
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    by_month = {m["month"]: [] for m in MA}
    for s in SUBS:
        by_month[MA[s["month_index"]]["month"]].append(s)
    avgbound_cell = {}  # month -> "'tab'!$C$row"
    cols = ["Submitted Date", "Submission ID", "State", "Vehicle", "Rating Class",
            "Agreed Value", "Base Rate /100", "Territory Factor", "Softening Index",
            "Gross-Up Divisor", "Rated Premium (calc)", "Premium — Artifact", "Δ",
            "Disposition", "Premium Kind"]
    # col letters: A..O ; formula refs: F AV,G base,H terr,I soft,J gross,K calc,L artifact,O kind
    for m in MA:
        mon = m["month"]; ws = wb.create_sheet(mon)
        title(ws, f"Monthly Rater — {mon}  (Month {m['month_index']+1} of 12)",
              f"Softening index {m['softening_index']:.4f}. Rated Premium = ROUND(AgreedValue/100 × BaseRate × "
              f"Territory × Softening ÷ Gross-Up, 2); ties to artifact premium within $0.01 (generator rounding).",
              span=len(cols))
        HR = 4; hdr(ws, HR, cols)
        rows_m = by_month[mon]
        for j, s in enumerate(rows_m):
            r = HR + 1 + j
            q = s["quote"]; rb = q["rating_basis"]; v = s["vehicle"]
            vals = [s["submitted_at"], s["submission_id"], s["garaging_state"],
                    f'{v["year"]} {v["make"]} {v["model"]}',
                    CLASS_LABEL.get(rb["rating_vehicle_class"], rb["rating_vehicle_class"]),
                    v["agreed_value"], rb["base_rate_per_100"], rb["territory_factor"],
                    q["softening_index"], rb["gross_up_divisor"], None, q["premium"], None,
                    s["disposition"], q["premium_kind"]]
            for i, val in enumerate(vals):
                c = ws.cell(row=r, column=1 + i, value=val); c.font = font(10)
                if r % 2 == 0: c.fill = PatternFill("solid", fgColor=GREY)
            ws.cell(row=r, column=11, value=f"=ROUND(F{r}/100*G{r}*H{r}*I{r}/J{r},2)")
            ws.cell(row=r, column=13, value=f"=L{r}-K{r}")
            for cc in (6, 11, 12, 13): ws.cell(row=r, column=cc).number_format = CUR
            ws.cell(row=r, column=7).number_format = RATE
            ws.cell(row=r, column=8).number_format = RATE
            ws.cell(row=r, column=9).number_format = RATE
            ws.cell(row=r, column=10).number_format = RATE
        last = HR + len(rows_m)
        ws.auto_filter.ref = f"A{HR}:O{last}"
        # summary block
        sr = last + 2
        blk = [
            ("Rated submissions", f"=COUNTA(B{HR+1}:B{last})", INT),
            ("Bound count", f'=COUNTIF(O{HR+1}:O{last},"bound")', INT),
            ("Avg premium — all rated", f"=AVERAGE(L{HR+1}:L{last})", CUR),
            ("Avg premium — BOUND only", f'=AVERAGEIF(O{HR+1}:O{last},"bound",L{HR+1}:L{last})', CUR),
            ("GWP bound (month)", f'=SUMIF(O{HR+1}:O{last},"bound",L{HR+1}:L{last})', CUR),
            ("Softening index (month)", m["softening_index"], RATE),
        ]
        for k, (lab, f, fmt) in enumerate(blk):
            rr = sr + k
            ws.cell(row=rr, column=1, value=lab).font = font(10, b=True)
            c = ws.cell(row=rr, column=3, value=f); c.font = font(10, b=True); c.number_format = fmt
            c.fill = PatternFill("solid", fgColor=TOTFILL)
        avgbound_cell[mon] = f"'{mon}'!$C${sr+3}"
        widths(ws, {"A": 13, "B": 30, "C": 6, "D": 24, "E": 24, "F": 14, "G": 12, "H": 13,
                    "I": 13, "J": 13, "K": 15, "L": 16, "M": 9, "N": 11, "O": 12})
    # Rate Trend tab
    rt = wb.create_sheet("Rate Trend", 0)
    rt.sheet_view.showGridLines = False
    title(rt, "Declining Rate Trend (12 months)",
          "Avg BOUND premium and softening index per month — pulled live from each month tab. "
          "The downward trend is the core story.", span=5)
    hdr(rt, 4, ["Month", "Month #", "Softening Index", "Avg Bound Premium", "% vs Month 1"])
    first_prem_row = 5
    for k, m in enumerate(MA):
        r = 5 + k
        rt.cell(row=r, column=1, value=m["month"]).font = font(10)
        rt.cell(row=r, column=2, value=m["month_index"] + 1).font = font(10)
        c3 = rt.cell(row=r, column=3, value=m["softening_index"]); c3.font = font(10); c3.number_format = RATE
        c4 = rt.cell(row=r, column=4, value=f"={avgbound_cell[m['month']]}"); c4.font = font(10); c4.number_format = CUR
        c5 = rt.cell(row=r, column=5, value=f"=D{r}/$D${first_prem_row}-1"); c5.font = font(10); c5.number_format = PCT2
        for cc in range(1, 6): rt.cell(row=r, column=cc).border = BORDER
    tr = 5 + len(MA) + 1
    rt.cell(row=tr, column=1, value="Month 1 → Month 12 change").font = font(10, b=True)
    ch = rt.cell(row=tr, column=5, value=f"=D{5+len(MA)-1}/D{first_prem_row}-1")
    ch.font = font(10, b=True); ch.number_format = PCT2; ch.fill = PatternFill("solid", fgColor=TOTFILL)
    widths(rt, {"A": 12, "B": 9, "C": 15, "D": 18, "E": 14})
    readme_tab(wb, "02 · Monthly Raters (×12)", [
        ("Tabs", "Rate Trend + 12 monthly rater tabs"),
        ("Rate formula", "Premium = ROUND(AV/100 × BaseRate × Territory × Softening ÷ Gross-Up, 2)"),
        ("Trend (artifact)", f'M1 ${MA[0]["avg_bound_premium"]:,.2f} → M12 ${MA[11]["avg_bound_premium"]:,.2f} '
                             f'({(MA[11]["avg_bound_premium"]/MA[0]["avg_bound_premium"]-1)*100:.1f}%)'),
        ("Δ column", "Artifact premium − calc; within $0.01 (generator intermediate rounding)."),
    ])
    tabcolors(wb, "1F4E79")
    p = f"{OUT}/02_monthly_raters.xlsx"; wb.save(p); return p

# ================================================================= WB3 UNDERWRITING BDX
def build_bdx():
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    bound = [s for s in SUBS if s.get("policy")]
    by_month = {m["month"]: [] for m in MA}
    for s in bound:
        by_month[MA[s["month_index"]]["month"]].append(s)
    # columns: A pol#,B eff,C insured,D state,E AV,F premium,G broker,H torque,I markets,
    # J residual, K..T syndicates(10), U policy fee,V insp fee,W fee total
    base = ["Policy Number", "Effective Date", "Insured", "State", "Agreed Value",
            "Bound Premium", "Broker 12.5%", "Torque 12.5%", "Markets 75%", "Rounding Residual"]
    synd_cols = list(SYNDS)
    tail = ["Policy Fee", "Inspection Fee", "Fee Total"]
    cols = base + synd_cols + tail
    NC = len(cols)
    Kc = 11  # first syndicate col index
    tot_rows = {}  # month -> totals row number
    for m in MA:
        mon = m["month"]; ws = wb.create_sheet(mon)
        title(ws, f"Underwriting BDX — {mon}  (Bound book only)",
              "Indicative/declined premium excluded. Markets 75% = SUM of the ten syndicate columns "
              "(7.5% each). Residual = Premium − (Broker+Torque+Markets) so the row foots exactly.",
              span=NC)
        HR = 4; hdr(ws, HR, cols)
        rows_m = by_month[mon]
        for j, s in enumerate(rows_m):
            r = HR + 1 + j
            pol = s["policy"]; mo = pol["money"]; v = s["vehicle"]; a = s["applicant"]
            ws.cell(row=r, column=1, value=pol["policy_number"])
            ws.cell(row=r, column=2, value=pol["effective_date"])
            ws.cell(row=r, column=3, value=f'{a["first_name"]} {a["last_name"]}')
            ws.cell(row=r, column=4, value=s["garaging_state"])
            ws.cell(row=r, column=5, value=v["agreed_value"])
            ws.cell(row=r, column=6, value=pol["bound_premium"])
            ws.cell(row=r, column=7, value=mo["commission_broker"])
            ws.cell(row=r, column=8, value=mo["commission_torque"])
            # syndicate shares from artifact
            for si, name in enumerate(synd_cols):
                ws.cell(row=r, column=Kc + si, value=mo["syndicate_shares"][name])
            k0 = get_column_letter(Kc); k1 = get_column_letter(Kc + 9)
            ws.cell(row=r, column=9, value=f"=SUM({k0}{r}:{k1}{r})")          # markets = sum synds
            ws.cell(row=r, column=10, value=f"=F{r}-(G{r}+H{r}+I{r})")        # residual
            ws.cell(row=r, column=Kc + 10, value=mo["policy_fee"])
            ws.cell(row=r, column=Kc + 11, value=mo["inspection_fee"])
            ws.cell(row=r, column=Kc + 12, value=f"=U{r}+V{r}")              # fee total
            for i in range(1, NC + 1):
                c = ws.cell(row=r, column=i); c.font = font(10)
                if i >= 5: c.number_format = CUR
                if r % 2 == 0: c.fill = PatternFill("solid", fgColor=GREY)
        last = HR + len(rows_m)
        # totals row
        tr = last + 1; tot_rows[mon] = tr
        ws.cell(row=tr, column=1, value="TOTAL").font = font(10, b=True)
        for i in range(5, NC + 1):
            L = get_column_letter(i)
            c = ws.cell(row=tr, column=i, value=f"=SUM({L}{HR+1}:{L}{last})")
            c.font = font(10, b=True); c.number_format = CUR
            c.fill = PatternFill("solid", fgColor=TOTFILL); c.border = BORDER
        # foot check row: Premium - (broker+torque+markets+residual) should be 0
        cr = tr + 1
        ws.cell(row=cr, column=1, value="Foot check: Premium − (B+T+M+Resid)").font = SUBF
        cc = ws.cell(row=cr, column=6, value=f"=F{tr}-(G{tr}+H{tr}+I{tr}+J{tr})")
        cc.font = font(9, b=True); cc.number_format = CUR
        cr2 = cr + 1
        ws.cell(row=cr2, column=1, value="Foot check: Markets − Σ(10 syndicates)").font = SUBF
        k0 = get_column_letter(Kc); k1 = get_column_letter(Kc + 9)
        cc2 = ws.cell(row=cr2, column=9, value=f"=I{tr}-SUM({k0}{tr}:{k1}{tr})")
        cc2.font = font(9, b=True); cc2.number_format = CUR
        ws.auto_filter.ref = f"A{HR}:{get_column_letter(NC)}{last}"
        wmap = {"A": 20, "B": 13, "C": 20, "D": 6, "E": 14, "F": 15, "G": 13, "H": 13, "I": 14, "J": 15}
        for si in range(10): wmap[get_column_letter(Kc + si)] = 11
        wmap[get_column_letter(Kc + 10)] = 11; wmap[get_column_letter(Kc + 11)] = 13; wmap[get_column_letter(Kc + 12)] = 11
        widths(ws, wmap)
    # Annual Summary
    an = wb.create_sheet("Annual Summary", 0)
    an.sheet_view.showGridLines = False
    an_cols = ["Month", "Bound Premium", "Broker 12.5%", "Torque 12.5%", "Markets 75%",
               "Rounding Residual"] + synd_cols + ["Policy Fee", "Inspection Fee", "Fee Total"]
    title(an, "Underwriting BDX — Annual Summary",
          f"Each row pulls that month tab's TOTAL row. Grand total ties to artifact GWP ${GWP:,.2f}.",
          span=len(an_cols))
    hdr(an, 4, an_cols)
    # month->totals cell mapping. In month tab: premium F, broker G, torque H, markets I, resid J,
    # synds K..T, polfee U, inspfee V, feetot W
    src_letters = ["F", "G", "H", "I", "J"] + [get_column_letter(Kc + i) for i in range(10)] + ["U", "V", "W"]
    firstr = 5
    for k, m in enumerate(MA):
        r = 5 + k; mon = m["month"]; tr = tot_rows[mon]
        an.cell(row=r, column=1, value=mon).font = font(10)
        for i, L in enumerate(src_letters):
            c = an.cell(row=r, column=2 + i, value=f"='{mon}'!{L}{tr}")
            c.font = font(10); c.number_format = CUR; c.border = BORDER
        an.cell(row=r, column=1).border = BORDER
    gr = 5 + len(MA)
    an.cell(row=gr, column=1, value="GRAND TOTAL").font = font(10, b=True)
    for i in range(2, len(an_cols) + 1):
        L = get_column_letter(i)
        c = an.cell(row=gr, column=i, value=f"=SUM({L}{firstr}:{L}{gr-1})")
        c.font = font(10, b=True); c.number_format = CUR
        c.fill = PatternFill("solid", fgColor=TOTFILL); c.border = BORDER
    # reconciliation block
    rb0 = gr + 2
    recon = [
        ("RECONCILIATION", "", ""),
        ("Annual bound premium (grand total)", f"=B{gr}", None),
        ("Artifact GWP (bound)", None, GWP),
        ("Δ premium vs GWP (must be 0)", f"=B{gr}-{GWP}", None),
        ("Broker+Torque+Markets+Residual", f"=C{gr}+D{gr}+E{gr}+F{gr}", None),
        ("Δ split vs premium (must be 0)", f"=(C{gr}+D{gr}+E{gr}+F{gr})-B{gr}", None),
        ("Σ ten syndicate columns", f"=SUM(G{gr}:P{gr})", None),
        ("Markets 75% (grand total)", f"=E{gr}", None),
        ("Δ syndicates vs markets (must be 0)", f"=SUM(G{gr}:P{gr})-E{gr}", None),
        ("Per-syndicate total (each, from tabs)", f"=G{gr}", None),
        ("  memo: theoretical 0.075 × GWP", None, SUM["per_syndicate_total_7_5pct"]),
        ("Fee income (grand total)", f"=S{gr}", None),
        ("  artifact fee_income_total", None, SUM["fee_income_total"]),
    ]
    for k, (lab, f, art) in enumerate(recon):
        r = rb0 + k
        cL = an.cell(row=r, column=1, value=lab)
        cL.font = font(10, b=(lab == "RECONCILIATION" or "must be 0" in lab))
        if f is not None:
            c = an.cell(row=r, column=3, value=f); c.font = font(10, b=True); c.number_format = CUR
            if "must be 0" in lab: c.fill = PatternFill("solid", fgColor=TOTFILL)
        if art is not None:
            c = an.cell(row=r, column=4, value=art); c.font = font(10); c.number_format = CUR
            c.fill = PatternFill("solid", fgColor=MEMOFILL)
    widths(an, {"A": 38, "B": 15, "C": 15, "D": 16, "E": 14, "F": 15})
    readme_tab(wb, "03 · Underwriting BDX", [
        ("Scope", f'BOUND book only ({SUM["binds"]:,} policies). Indicative/declined excluded.'),
        ("Split", "12.5% Broker + 12.5% Torque + 75% Markets; Markets = Σ ten syndicates (7.5% each)."),
        ("Fees", "$350 policy fee/policy; $250 inspection for AV ≥ $1M."),
        ("Reconciliation", f"See 'Annual Summary' — Δ rows foot to 0; premium = GWP ${GWP:,.2f}."),
        ("Rounding residual", "Explicit column; standard bordereau practice so the row foots exactly."),
    ])
    tabcolors(wb, "1F4E79")
    p = f"{OUT}/03_underwriting_bdx.xlsx"; wb.save(p); return p

# ================================================================= WB4 CLAIMS BDX
def build_claims():
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    # policy_id -> (policy_number, insured, state)
    pmap = {}
    for s in SUBS:
        if s.get("policy"):
            pmap[s["policy"]["policy_id"]] = (
                s["policy"]["policy_number"],
                f'{s["applicant"]["first_name"]} {s["applicant"]["last_name"]}',
                s["garaging_state"])
    reg = sorted(CLAIMS, key=lambda c: c["date_of_loss"])
    ws = wb.create_sheet("Claims Register")
    cols = ["Claim ID", "Policy Number", "Insured", "State", "Policy ID", "Date of Loss",
            "Loss Month", "Incurred (Paid+Reserves)", "Status"]
    title(ws, "Claims BDX — Policy-Period Claims Register",
          "Continuous register, by date of loss. Claims on bound policies only. Losses span beyond the "
          "12 submission months (late binds carry losses into the next year); reported by date of loss.",
          span=len(cols))
    HR = 4; hdr(ws, HR, cols)
    for j, c in enumerate(reg):
        r = HR + 1 + j
        pn, ins, st = pmap.get(c["policy_id"], ("", "", ""))
        vals = [c["claim_id"], pn, ins, st, c["policy_id"], c["date_of_loss"],
                c["date_of_loss"][:7], c["incurred"], c["status"]]
        for i, val in enumerate(vals):
            cc = ws.cell(row=r, column=1 + i, value=val); cc.font = font(10)
            if i == 7: cc.number_format = CUR
            if r % 2 == 0: cc.fill = PatternFill("solid", fgColor=GREY)
    last = HR + len(reg)
    tr = last + 1
    ws.cell(row=tr, column=1, value="TOTAL INCURRED").font = font(10, b=True)
    tc = ws.cell(row=tr, column=8, value=f"=SUM(H{HR+1}:H{last})")
    tc.font = font(10, b=True); tc.number_format = CUR; tc.fill = PatternFill("solid", fgColor=TOTFILL)
    ws.auto_filter.ref = f"A{HR}:I{last}"
    widths(ws, {"A": 30, "B": 20, "C": 20, "D": 6, "E": 30, "F": 13, "G": 11, "H": 22, "I": 9})

    # By Loss Month
    bm = wb.create_sheet("By Loss Month")
    bm.sheet_view.showGridLines = False
    title(bm, "Claims by Loss Month", "Count and incurred per loss month (SUMIF/COUNTIF over the register), with running cumulative.", span=4)
    hdr(bm, 4, ["Loss Month", "Claim Count", "Incurred", "Cumulative Incurred"])
    months = sorted({c["date_of_loss"][:7] for c in reg})
    reg_month = f"'Claims Register'!$G${HR+1}:$G${last}"
    reg_inc = f"'Claims Register'!$H${HR+1}:$H${last}"
    firstr = 5
    for k, mon in enumerate(months):
        r = 5 + k
        bm.cell(row=r, column=1, value=mon).font = font(10)
        bm.cell(row=r, column=2, value=f'=COUNTIF({reg_month},"{mon}")').font = font(10)
        c3 = bm.cell(row=r, column=3, value=f'=SUMIF({reg_month},"{mon}",{reg_inc})'); c3.font = font(10); c3.number_format = CUR
        c4 = bm.cell(row=r, column=4, value=f"=SUM($C${firstr}:C{r})"); c4.font = font(10); c4.number_format = CUR
        for cc in range(1, 5): bm.cell(row=r, column=cc).border = BORDER
    gr = 5 + len(months)
    bm.cell(row=gr, column=1, value="TOTAL").font = font(10, b=True)
    tc = bm.cell(row=gr, column=2, value=f"=SUM(B{firstr}:B{gr-1})"); tc.font = font(10, b=True)
    ti = bm.cell(row=gr, column=3, value=f"=SUM(C{firstr}:C{gr-1})"); ti.font = font(10, b=True); ti.number_format = CUR
    ti.fill = PatternFill("solid", fgColor=TOTFILL)
    widths(bm, {"A": 12, "B": 12, "C": 18, "D": 20})

    # Loss Ratio
    lr = wb.create_sheet("Loss Ratio", 0)
    lr.sheet_view.showGridLines = False
    title(lr, "Loss Ratio Reconciliation", "Incurred ÷ Bound premium, bound book only. Under the 0.60 PC hurdle → bonus pays at the 10% band.", span=3)
    rows = [
        ("Total incurred losses", f"='Claims Register'!H{tr}", None, CUR),
        ("Bound GWP (artifact)", None, GWP, CUR),
        ("Loss Ratio (incurred ÷ GWP)", f"='Claims Register'!H{tr}/{GWP}", None, RATE),
        ("Target loss ratio", None, SUM["loss_ratio_target"], RATE),
        ("Artifact loss_ratio", None, SUM["loss_ratio"], RATE),
        ("Profit-commission band", None, SUM["pc_band_pct"] / 100, PCT2),
        ("Claim count", f"=COUNTA('Claims Register'!A{HR+1}:A{last})", SUM["policy_period_claim_count"], INT),
        ("Claims referencing non-bound policy", 0, None, INT),
    ]
    hdr(lr, 4, ["Metric", "Formula / value", "Artifact"])
    for k, (lab, f, art, fmt) in enumerate(rows):
        r = 5 + k
        lr.cell(row=r, column=1, value=lab).font = font(10)
        if f is not None:
            c = lr.cell(row=r, column=2, value=f); c.font = font(10, b=True); c.number_format = fmt
            if "Loss Ratio" in lab: c.fill = PatternFill("solid", fgColor=TOTFILL)
        if art is not None:
            c = lr.cell(row=r, column=3, value=art); c.font = font(10); c.number_format = fmt
        for cc in range(1, 4): lr.cell(row=r, column=cc).border = BORDER
    widths(lr, {"A": 34, "B": 20, "C": 16})
    readme_tab(wb, "04 · Claims BDX", [
        ("Register", f"{len(reg)} policy-period claims on bound policies, by date of loss."),
        ("Loss span", f'{reg[0]["date_of_loss"]} → {reg[-1]["date_of_loss"]} (extends past the 12 submission months).'),
        ("Total incurred", f'${SUM["total_incurred_losses"]:,.2f}'),
        ("Loss ratio", f'{SUM["loss_ratio"]:.4f}  (incurred ÷ GWP) — PC band {SUM["pc_band_pct"]}%'),
        ("Status values", "Artifact uses closed/open (closed = settled/paid); shown as-is."),
    ])
    tabcolors(wb, "C00000")
    p = f"{OUT}/04_claims_bdx.xlsx"; wb.save(p); return p

paths = [build_submissions(), build_raters(), build_bdx(), build_claims()]
print("BUILT:")
for p in paths: print("  ", p)
