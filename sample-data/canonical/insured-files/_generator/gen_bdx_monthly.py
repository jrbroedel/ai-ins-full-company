#!/usr/bin/env python3
"""Monthly PROGRAM bordereaux: one file per month, listing every certificate bound that
month across the book (Kent's model). Technicals for the three multi-vehicle schedules come
from each rater's Portfolio Register (rater <-> BDX tie-out); Fairweather's and Blackwood's
rows and headers are copied verbatim from their shipped single-row files."""
import json, os, sys, calendar, subprocess, datetime as dt
import openpyxl

SRC = "bdx_template_repaired.xlsx"
RECALC = "/mnt/skills/public/xlsx/scripts/recalc.py"
OUT = "out/bdx_monthly"; os.makedirs(OUT, exist_ok=True)
WS_SHIPPED = "/home/claude/ws/01_Completed_Files"
INPUT_COLS = ["A","B","C","D","E","F","H","I","J","K","L","M","N","O","P","Q","R","S",
              "U","V","W","X","AA","AB","AC","AE","AF","AG","AL","BA","BC"]
STATEFULL = {"KS":"Kansas","CA":"California","MA":"Massachusetts"}
ORIG = {"Delacroix": 147, "Vasquez": 148, "Harrington": 149}

def sched_rows(name):
    spec = json.load(open(f"spec_{name}.json"))
    reg = json.load(open("register_manifest.json"))[name]
    certs = json.load(open("cert_map.json"))
    st = spec["state"]
    master = f"B0999TQ26{ORIG[name]:04d}"
    rows = []
    for vi, v in enumerate(spec["vehicles"]):
        r = reg[str(vi + 1)]
        cert = certs[f"{name},{vi}"]
        eff = dt.date.fromisoformat(v["effective_date"])
        rows.append({"A": f"TQ-C-2026-{cert:04d}", "B": spec["insured"]["name"], "C": eff,
            "D": "Bound - within authority", "E": "New business", "F": f"B0999TQ26{cert:04d}",
            "H": int(v["year"]), "I": v.get("rater_make", v["make"]), "J": v["model"],
            "K": v["vin"], "L": v["vclass"], "M": v["agreed_value"], "N": "Personal",
            "O": "Personal - regular use", "P": "1,000 - 2,500", "Q": STATEFULL[st],
            "R": v["garaging_zip"], "S": r["zone"],
            "U": "2% of agreed value, minimum $10,000", "V": "Agreed Value",
            "W": v["pd_deductible_row"], "X": "$1,000,000 CSL", "AA": "Surplus Lines",
            "AB": round(r["technical"], 2), "AC": 0, "AE": 250,
            "AF": 150 if v["agreed_value"] >= 250000 else 0, "AG": spec["sl_rate"],
            "AL": "At technical", "BA": "Yes",
            "BC": f"Register ref {v['policy_number']}. Bound within delegated authority. "
                  f"Schedule cert under master UMR {master}. SL tax rate verified {STATEFULL[st]}."})
    eff0 = dt.date.fromisoformat(spec["vehicles"][0]["effective_date"])
    return (eff0.year, eff0.month), rows, master

def shipped_rows(folder, fname):
    wb = openpyxl.load_workbook(f"{WS_SHIPPED}/{folder}/{fname}")
    ws = wb["UW Bordereaux"]
    row = {c: ws[f"{c}5"].value for c in INPUT_COLS}
    hdr = {f"C{i}": wb["Header"][f"C{i}"].value for i in range(5, 21)}
    return row, hdr

def main():
    months = {}
    headers = {}
    for name in ["Delacroix", "Vasquez", "Harrington"]:
        (yy, mm), rows, master = sched_rows(name)
        months.setdefault((yy, mm), []).extend(rows)
        last = calendar.monthrange(yy, mm)[1]
        headers[(yy, mm)] = {"C5": master, "C6": "BA-EXO-2026-01", "C7": "Torque Underwriters LLC",
            "C8": "TQ-COV-0000", "C9": "MT", "C10": yy,
            "C11": "Exotic, supercar and collector automobile",
            "C12": dt.date(yy, mm, 1), "C13": dt.date(yy, mm, last), "C14": dt.date(yy, mm, last),
            "C15": "Torque Underwriters, Delegated Authority", "C16": "USD", "C17": "USD",
            "C18": 2000000, "C19": -0.15, "C20": 5000}
    for folder, fname, ym in [("Fairweather, Noor", "03_BDX_2025-11_Fairweather_Noor.xlsx", (2025, 11)),
                              ("Blackwood, Cassius", "03_BDX_2025-12_Blackwood_Cassius.xlsx", (2025, 12))]:
        row, hdr = shipped_rows(folder, fname)
        months.setdefault(ym, []).append(row)
        headers[ym] = hdr
    for (yy, mm), rows in sorted(months.items()):
        wb = openpyxl.load_workbook(SRC)
        ws = wb["UW Bordereaux"]
        r = 5
        for vals in rows:
            for col, val in vals.items(): ws[f"{col}{r}"] = val
            r += 1
        for rr in range(r, 11):
            for col in INPUT_COLS: ws[f"{col}{rr}"] = None
        for k, v in headers[(yy, mm)].items(): wb["Header"][k] = v
        outp = f"{OUT}/03_BDX_{yy:04d}-{mm:02d}.xlsx"
        wb.save(outp)
        subprocess.run([sys.executable, RECALC, outp], capture_output=True, text=True)
        wv = openpyxl.load_workbook(outp, data_only=True)["UW Bordereaux"]
        rr = 5; ok = True; tot = 0.0
        while wv[f"A{rr}"].value:
            ok &= (wv[f"AX{rr}"].value == "PASS"); tot += wv[f"AB{rr}"].value or 0; rr += 1
        print(f"{yy:04d}-{mm:02d}: {rr-5} rows, allPASS={ok}, technical sum {tot:,.2f}")

if __name__ == "__main__":
    main()
