#!/usr/bin/env python3
"""Rebuild each client's single rater with the Portfolio Register carrying the FULL
schedule (one row per vehicle, cert-referenced), specimen rows cleared, Rating Engine
still holding the lead vehicle. Raw-XML patch of sheets 3/4/5/6 preserves x14 dropdowns
and charts; recalced caches come from an openpyxl working copy (shipped discipline)."""
import json, os, re, shutil, subprocess, sys, zipfile, datetime as dt
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from gen_raters import SRC, RECALC, esc, serial, inputs_for, read_vals, patch_cell as _patch_cell

def patch_cell(xml, ref, value, dtype, keep_formula=True):
    if isinstance(value, (dt.datetime, dt.date)):
        d0 = value.date() if isinstance(value, dt.datetime) else value
        value = (d0 - dt.date(1899, 12, 30)).days
        dtype = "n"
    return _patch_cell(xml, ref, value, dtype, keep_formula)

OUTDIR = "out/raters_v2"; os.makedirs(OUTDIR, exist_ok=True)
ENDOR = [get_column_letter(c) for c in range(28, 43)]  # AB..AP -> "None"

def reg_row(spec, v, cert):
    eff = dt.date.fromisoformat(v["effective_date"])
    exp = eff.replace(year=eff.year + 1)
    d = {"A": f"TQ-C-2026-{cert:04d}", "B": spec["insured"]["name"],
         "C": ("d", serial(v["effective_date"])), "D": ("d", (exp - dt.date(1899,12,30)).days),
         "E": int(v["year"]), "F": v.get("rater_make", v["make"]), "G": v["model"],
         "H": v["vin"], "I": v["vclass"], "J": v["agreed_value"],
         "K": spec["state_full"], "L": v["garaging_zip"], "M": "Personal",
         "N": "Personal - regular use", "O": "1,000 - 2,500",
         "P": "Locked private garage w/ monitored alarm",
         "Q": "GPS tracker w/ recovery service",
         "R": spec["policy_selections"]["age_band"], "S": "5 - 10 years",
         "T": "Clean - no violations 3 yrs", "U": "No at-fault claims - 5 yrs",
         "V": spec["policy_selections"]["collection_size"], "W": v["pd_deductible_row"],
         "X": "Agreed Value", "Y": "$1,000,000 CSL", "Z": spec["metro"],
         "AA": "2% of agreed value, minimum $10,000", "AQ": 0}
    return d

def row_styles(xml):
    """col letter -> s= attr from row 5 (specimen) for style borrowing."""
    styles = {}
    m = re.search(r'<row r="5"[^>]*>(.*?)</row>', xml, re.S)
    if m:
        for cm in re.finditer(r'<c r="([A-Z]+)5"([^>]*?)(?:/>|>)', m.group(1)):
            sm = re.search(r'\bs="(\d+)"', cm.group(2))
            if sm: styles[cm.group(1)] = sm.group(1)
    return styles

def upsert_cell(xml, ref, value, styles, dtype="auto"):
    """Patch existing cell or insert into its <row> in column order."""
    col = re.match(r'([A-Z]+)', ref).group(1); rown = ref[len(col):]
    if re.search(r'<c r="' + re.escape(ref) + r'"', xml):
        if value is None or value == "":
            # clear: keep style, drop content
            pat = re.compile(r'<c r="' + re.escape(ref) + r'"([^>]*?)(?:/>|>.*?</c>)', re.S)
            m = pat.search(xml)
            sm = re.search(r'\bs="(\d+)"', m.group(1)); s = f' s="{sm.group(1)}"' if sm else ''
            return xml[:m.start()] + f'<c r="{ref}"{s}/>' + xml[m.end():]
        dt_ = "n" if isinstance(value, (int, float)) else "s"
        if dtype == "d": dt_ = "n"
        return patch_cell(xml, ref, value, dt_, keep_formula=False)
    if value is None or value == "": return xml
    s = f' s="{styles[col]}"' if col in styles else ''
    if isinstance(value, str):
        new = f'<c r="{ref}"{s} t="inlineStr"><is><t xml:space="preserve">{esc(value)}</t></is></c>'
    else:
        new = f'<c r="{ref}"{s}><v>{value}</v></c>'
    rm = re.search(r'<row r="' + rown + r'"[^>]*>(.*?)</row>', xml, re.S)
    if not rm:
        # row absent entirely: insert row before the next existing row
        nxt = re.search(r'<row r="(\d+)"', xml[rm.end():] if rm else xml)
        raise RuntimeError(f"row {rown} absent in sheet XML")
    inner = rm.group(1); target = column_index_from_string(col)
    pos = len(inner)
    for cm in re.finditer(r'<c r="([A-Z]+)' + rown + r'"', inner):
        if column_index_from_string(cm.group(1)) > target:
            pos = cm.start(); break
    inner2 = inner[:pos] + new + inner[pos:]
    return xml[:rm.start(1)] + inner2 + xml[rm.end(1):]

def build_client(name, certs):
    spec = json.load(open(f"spec_{name}.json"))
    n = len(spec["vehicles"])
    work = "/tmp/_w_reg.xlsx"; shutil.copy(SRC, work)
    wb = openpyxl.load_workbook(work)
    reE = wb["Rating Engine"]
    inp = inputs_for(spec, spec["vehicles"][0])
    for cell, val, dtp in inp:
        reE[cell] = (dt.date(1899,12,30)+dt.timedelta(days=val)) if dtp == "d" else val
    reE["C44"] = 0
    reg = wb["Portfolio Register"]
    for vi, v in enumerate(spec["vehicles"]):
        rr = 5 + vi
        for col, val in reg_row(spec, v, certs[f"{name},{vi}"]).items():
            if isinstance(val, tuple):
                reg[f"{col}{rr}"] = dt.date(1899,12,30)+dt.timedelta(days=val[1])
            else:
                reg[f"{col}{rr}"] = val
    for rr in range(5 + n, 11):  # clear leftover specimen inputs
        for c in range(1, 44): reg.cell(row=rr, column=c).value = None
    wb.save(work)
    subprocess.run([sys.executable, RECALC, work], capture_output=True, text=True)
    reV = read_vals(work, "Rating Engine"); pbV = read_vals(work, "Premium Breakout")
    rgV = read_vals(work, "Portfolio Register"); psV = read_vals(work, "Portfolio Summary")

    z = zipfile.ZipFile(SRC); parts = {p: z.read(p) for p in z.namelist()}; z.close()
    s3 = parts["xl/worksheets/sheet3.xml"].decode()
    s4 = parts["xl/worksheets/sheet4.xml"].decode()
    s5 = parts["xl/worksheets/sheet5.xml"].decode()
    s6 = parts["xl/worksheets/sheet6.xml"].decode()
    for cell, val, dtp in inp: s3 = patch_cell(s3, cell, val, dtp, keep_formula=False)
    s3 = patch_cell(s3, "C44", 0, "n", keep_formula=False)
    for coord, val in reV.items():
        if coord[0] == "C" and coord[1:].isdigit() and 45 <= int(coord[1:]) <= 123:
            s3 = patch_cell(s3, coord, val, "n", keep_formula=True)
    for coord, val in pbV.items(): s4 = patch_cell(s4, coord, val, "n", keep_formula=True)

    styles = row_styles(s5)
    for vi, v in enumerate(spec["vehicles"]):
        rr = 5 + vi
        for col, val in reg_row(spec, v, certs[f"{name},{vi}"]).items():
            if isinstance(val, tuple): s5 = upsert_cell(s5, f"{col}{rr}", val[1], styles, dtype="d")
            else: s5 = upsert_cell(s5, f"{col}{rr}", val, styles)
    for rr in range(5 + n, 11):
        for c in range(1, 44):
            s5 = upsert_cell(s5, f"{get_column_letter(c)}{rr}", None, styles)
    # refresh register + summary formula caches
    for coord, val in rgV.items():
        cl = re.match(r'([A-Z]+)(\d+)', coord)
        if cl and column_index_from_string(cl.group(1)) >= 44:
            s5 = patch_cell(s5, coord, val, "n", keep_formula=True)
    # formulas that recalced to empty (cleared rows) need their stale caches emptied
    for rr in range(5 + n, 11):
        for c in range(44, 105):
            coord = f"{get_column_letter(c)}{rr}"
            if coord not in rgV: s5 = patch_cell(s5, coord, "", "s", keep_formula=True)
    for coord, val in psV.items(): s6 = patch_cell(s6, coord, val, "n", keep_formula=True)

    parts["xl/worksheets/sheet3.xml"] = s3.encode()
    parts["xl/worksheets/sheet4.xml"] = s4.encode()
    parts["xl/worksheets/sheet5.xml"] = s5.encode()
    parts["xl/worksheets/sheet6.xml"] = s6.encode()
    wbx = parts["xl/workbook.xml"].decode()
    if "<calcPr" in wbx: wbx = re.sub(r'<calcPr[^>]*/>', '<calcPr calcId="191029" fullCalcOnLoad="1"/>', wbx)
    else: wbx = wbx.replace("</workbook>", '<calcPr calcId="191029" fullCalcOnLoad="1"/></workbook>')
    parts["xl/workbook.xml"] = wbx.encode()
    first = spec["insured"]["name"].split()[0]
    outp = f"{OUTDIR}/02_Rater_{name}_{first}.xlsx"
    with zipfile.ZipFile(outp, "w", zipfile.ZIP_DEFLATED) as zo:
        for p, data in parts.items(): zo.writestr(p, data)
    # report per-row register GWP vs technical manifest
    gwps = [rgV.get(f"CO{5+i}") for i in range(n)]
    return outp, gwps, rgV

if __name__ == "__main__":
    certs = json.load(open("cert_map.json"))
    rman = json.load(open("rater_manifest.json"))
    for name in ["Delacroix", "Vasquez", "Harrington"]:
        outp, gwps, rgV = build_client(name, certs)
        n = len(gwps)
        tot = sum(g for g in gwps if isinstance(g, (int, float)))
        print(f"{name}: {n} register rows, GWP col sums {tot:,.2f}; "
              f"rows: {[f'{g:,.0f}' if isinstance(g,(int,float)) else str(g) for g in gwps]}", flush=True)
