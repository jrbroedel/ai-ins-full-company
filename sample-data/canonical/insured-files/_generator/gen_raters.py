#!/usr/bin/env python3
"""Per-vehicle rater fill, generalized from the shipped _generator/fill_rater.py.
Same discipline: populate Rating Engine Section 1 on a working copy via openpyxl +
LibreOffice recalc for cached values, then patch the ORIGINAL workbook's sheet XML by
targeted regex (preserving s= styles, <f> formulas, x14 dropdowns, charts) and set
fullCalcOnLoad. Judgement adjustment C44 stays 0 -> rater shows technical premium."""
import json, os, re, shutil, subprocess, sys, zipfile, datetime as dt
import openpyxl

SRC = "/home/claude/repo/templates/Exotic_Collector_Auto_MGA_Rating_Matrix_7.xlsx"
RECALC = "/mnt/skills/public/xlsx/scripts/recalc.py"
OUTDIR = "out/raters"
os.makedirs(OUTDIR, exist_ok=True)

def esc(x): return str(x).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
def serial(dstr): return (dt.date.fromisoformat(dstr) - dt.date(1899,12,30)).days

def inputs_for(spec, v):
    ded = v["pd_deductible_row"]
    return [
        ("C5", spec["insured"]["name"], "s"),
        ("C6", serial(v["effective_date"]), "d"),
        ("C7", int(v["year"]), "n"), ("C8", v.get("rater_make", v["make"]), "s"), ("C9", v["model"], "s"),
        ("C10", v["vclass"], "s"), ("C11", v["agreed_value"], "n"),
        ("C12", spec["state_full"], "s"), ("C13", v["garaging_zip"], "s"),
        ("C14", "Personal", "s"), ("C15", "Personal - regular use", "s"),
        ("C16", "1,000 - 2,500", "s"),
        ("C17", "Locked private garage w/ monitored alarm", "s"),
        ("C18", "GPS tracker w/ recovery service", "s"),
        ("C19", spec["policy_selections"]["age_band"], "s"),
        ("C20", "5 - 10 years", "s"),
        ("C21", "Clean - no violations 3 yrs", "s"),
        ("C22", "No at-fault claims - 5 yrs", "s"),
        ("C23", spec["policy_selections"]["collection_size"], "s"),
        ("C24", ded, "s"), ("C25", "Agreed Value", "s"), ("C26", "$1,000,000 CSL", "s"),
        ("C27", spec["metro"], "s"),
        ("C28", "2% of agreed value, minimum $10,000", "s"),
    ]

def read_vals(path, sheet):
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    return {c.coordinate: c.value for row in ws.iter_rows() for c in row if c.value is not None}

def patch_cell(xml, ref, value, dtype, keep_formula=True):
    pat = re.compile(r'<c r="'+re.escape(ref)+r'"([^>]*?)(?:/>|>(.*?)</c>)', re.S)
    m = pat.search(xml)
    if not m: return xml
    attrs = m.group(1); inner = m.group(2) or ""
    sm = re.search(r'\bs="(\d+)"', attrs); s = f' s="{sm.group(1)}"' if sm else ''
    fm = re.search(r'<f[^>]*>.*?</f>|<f[^>]*/>', inner, re.S)
    if keep_formula and not fm: return xml
    if fm and keep_formula:
        f = fm.group(0)
        if isinstance(value, str): new = f'<c r="{ref}"{s} t="str">{f}<v>{esc(value)}</v></c>'
        else:
            val = value if not isinstance(value, bool) else int(value)
            new = f'<c r="{ref}"{s}>{f}<v>{val}</v></c>'
    else:
        if dtype == "s": new = f'<c r="{ref}"{s} t="inlineStr"><is><t xml:space="preserve">{esc(value)}</t></is></c>'
        else: new = f'<c r="{ref}"{s}><v>{value}</v></c>'
    return xml[:m.start()] + new + xml[m.end():]

def build(spec, v, out_path):
    inp = inputs_for(spec, v)
    work = "/tmp/_w_rater.xlsx"; shutil.copy(SRC, work)
    wb = openpyxl.load_workbook(work); re_ = wb["Rating Engine"]
    for cell, val, dtp in inp:
        re_[cell] = (dt.date(1899,12,30)+dt.timedelta(days=val)) if dtp == "d" else val
    re_["C44"] = 0
    wb.save(work)
    subprocess.run([sys.executable, RECALC, work], capture_output=True, text=True)
    reV = read_vals(work, "Rating Engine"); pbV = read_vals(work, "Premium Breakout")
    z = zipfile.ZipFile(SRC); parts = {n: z.read(n) for n in z.namelist()}; z.close()
    re_xml = parts["xl/worksheets/sheet3.xml"].decode()
    pb_xml = parts["xl/worksheets/sheet4.xml"].decode()
    for cell, val, dtp in inp: re_xml = patch_cell(re_xml, cell, val, dtp, keep_formula=False)
    re_xml = patch_cell(re_xml, "C44", 0, "n", keep_formula=False)
    for coord, val in reV.items():
        if coord[0] == "C" and coord[1:].isdigit() and 45 <= int(coord[1:]) <= 123:
            re_xml = patch_cell(re_xml, coord, val, "n", keep_formula=True)
    for coord, val in pbV.items():
        pb_xml = patch_cell(pb_xml, coord, val, "n", keep_formula=True)
    parts["xl/worksheets/sheet3.xml"] = re_xml.encode()
    parts["xl/worksheets/sheet4.xml"] = pb_xml.encode()
    wbx = parts["xl/workbook.xml"].decode()
    if "<calcPr" in wbx: wbx = re.sub(r'<calcPr[^>]*/>', '<calcPr calcId="191029" fullCalcOnLoad="1"/>', wbx)
    else: wbx = wbx.replace("</workbook>", '<calcPr calcId="191029" fullCalcOnLoad="1"/></workbook>')
    parts["xl/workbook.xml"] = wbx.encode()
    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zo:
        for n, data in parts.items(): zo.writestr(n, data)
    return dict(technical=reV.get("C105"), gwp=reV.get("C108"), status=reV.get("C122"),
                zone=reV.get("C48"), reason=reV.get("C123"))

if __name__ == "__main__":
    names = sys.argv[1:] or ["Delacroix", "Vasquez", "Harrington"]
    manifest = json.load(open("rater_manifest.json")) if os.path.exists("rater_manifest.json") else {}
    for name in names:
        spec = json.load(open(f"spec_{name}.json"))
        manifest.setdefault(name, {})
        for vi, v in enumerate(spec["vehicles"], 1):
            key = str(vi)
            if key in manifest[name]: continue
            out = f"{OUTDIR}/Rater_{name}_V{vi:02d}.xlsx"
            r = build(spec, v, out)
            r["frozen"] = v["bound_premium"]; r["seq"] = v["seq"]; r["file"] = out
            manifest[name][key] = r
            tech = r["technical"]
            techs = f"{tech:>10,.2f}" if isinstance(tech,(int,float)) else str(tech)
            print(f"{name} V{vi:02d} {v['make'][:14]:14} frozen={v['bound_premium']:>10,.2f} tech={techs} status={r['status']}", flush=True)
            json.dump(manifest, open("rater_manifest.json", "w"), indent=1)
    print("done")
